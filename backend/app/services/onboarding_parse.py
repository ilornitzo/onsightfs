from __future__ import annotations

import csv
import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


def _collapse_ws(text: str) -> str:
    return " ".join(text.split())


def _extract_text(pdf_bytes: bytes) -> tuple[str, list[str]]:
    reader = PdfReader(BytesIO(pdf_bytes))
    chunks: list[str] = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            chunks.append(page_text)

    full_text = "\n".join(chunks)
    lines = [_collapse_ws(line.strip()) for line in full_text.splitlines() if line and line.strip()]
    return full_text, lines


def _find_first(pattern: str, text: str, flags: int = re.IGNORECASE) -> str:
    match = re.search(pattern, text, flags)
    if not match:
        return ""
    return _collapse_ws(match.group(1).strip())


def _find_line_value(lines: list[str], label_pattern: str) -> str:
    regex = re.compile(label_pattern, re.IGNORECASE)
    for line in lines:
        match = regex.search(line)
        if match and match.group(1).strip():
            return _collapse_ws(match.group(1).strip())
    return ""


def _find_address(lines: list[str]) -> str:
    labeled = _find_line_value(lines, r"address\s*[:\-]\s*(.+)$")
    if labeled:
        return labeled

    street_like = re.compile(
        r"\b\d{1,6}\s+[A-Za-z0-9.\- ]+\s(?:st|street|ave|avenue|rd|road|dr|drive|ln|lane|blvd|boulevard|ct|court)\b",
        re.IGNORECASE,
    )
    for line in lines:
        if street_like.search(line):
            return line
    return ""


def _find_routing_account(lines: list[str], full_text: str) -> tuple[str, str]:
    routing = ""
    account = ""

    routing_line_re = re.compile(r"routing[^0-9]{0,30}(\d{9})", re.IGNORECASE)
    account_line_re = re.compile(r"account[^0-9]{0,30}(\d{4,17})", re.IGNORECASE)

    for line in lines:
        if not routing:
            match = routing_line_re.search(line)
            if match:
                routing = match.group(1)
        if not account:
            match = account_line_re.search(line)
            if match:
                account = match.group(1)
        if routing and account:
            break

    if not routing:
        routing = _find_first(r"routing(?:\s*number)?[^0-9]{0,40}(\d{9})", full_text)
    if not account:
        account = _find_first(r"account(?:\s*number)?[^0-9]{0,40}(\d{4,17})", full_text)

    return routing, account


def parse_onboarding_pdf(file_bytes: bytes) -> dict[str, Any]:
    full_text, lines = _extract_text(file_bytes)
    notes: list[str] = []

    name = _find_line_value(lines, r"name\s*[:\-]\s*(.+)$")
    business_name = _find_line_value(lines, r"business\s*name\s*[:\-]\s*(.+)$")

    if not name and lines:
        name = lines[0]
        notes.append("Name inferred from first non-empty line.")

    email = _find_first(r"\b([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})\b", full_text, flags=re.IGNORECASE)
    phone = _find_first(r"(\+?1?[\s.\-()]?\d{3}[\s.\-)]?\d{3}[\s.\-]?\d{4})", full_text, flags=re.IGNORECASE)
    dob = _find_first(r"\b(\d{2}/\d{2}/\d{4})\b", full_text, flags=re.IGNORECASE)
    ein_or_ssn = _find_first(r"\b(\d{2}-\d{7}|\d{3}-\d{2}-\d{4})\b", full_text, flags=re.IGNORECASE)

    address = _find_address(lines)
    bank_routing, bank_account = _find_routing_account(lines, full_text)

    if not business_name:
        maybe_business = _find_line_value(lines, r"(?:company|entity)\s*name\s*[:\-]\s*(.+)$")
        if maybe_business:
            business_name = maybe_business

    if not address:
        notes.append("Address not confidently detected.")
    if not bank_routing and not bank_account:
        notes.append("Bank routing/account not confidently detected.")

    return {
        "name": name or "",
        "business_name": business_name or "",
        "email": email or "",
        "phone": phone or "",
        "address": address or "",
        "dob": dob or "",
        "ein_or_ssn": ein_or_ssn or "",
        "bank_routing": bank_routing or "",
        "bank_account": bank_account or "",
        "notes": " ".join(notes).strip(),
    }


_TARGET_ALIASES: dict[str, list[str]] = {
    "name": ["full legal name", "legal name", "name"],
    "business_name": ["business name", "company name", "entity name"],
    "email": ["email address", "email"],
    "phone": ["phone number", "phone", "mobile", "cell"],
    "address": ["residential address", "address", "street address"],
    "dob": ["date of birth", "dob", "birth date"],
    "ssn_or_ein": ["ssn or ein", "ssn/ein", "ein/ssn", "ein", "ssn"],
    "aspen_grove_abc_number": ["aspen grove abc number", "abc number", "aspen grove", "aspen grove abc"],
    "bank_name": ["bank name", "financial institution", "name of bank"],
    "bank_account_type": ["account type", "bank account type", "checking or savings"],
    "bank_routing_number": ["routing number", "bank routing", "bank routing number"],
    "bank_account_number": ["account number", "bank account", "bank account number"],
    "counties": ["counties willing to cover", "counties", "service counties"],
    "expected_pay_per_inspection": ["expected pay per inspection", "pay per inspection"],
    "min_daily_volume": ["minimum daily volume", "min daily volume"],
    "ic_acknowledged": [
        "independent contractor acknowledgment",
        "independent contractor acknowledgement",
        "independent contractor acknowledgment accepted",
    ],
    "signature_name": ["digital signature", "signature", "signature name"],
    "signature_date": ["date signed", "signature date"],
    "notes": ["notes", "note", "comments"],
}


def _normalize_text(value: Any) -> str:
    text = _collapse_ws(str(value or "").strip().strip("'\"")).lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return _collapse_ws(text)


def _append_note(parsed: dict[str, str], message: str) -> None:
    existing = parsed.get("notes", "").strip()
    parsed["notes"] = f"{existing} {message}".strip()


def _append_detected_keys_note(parsed: dict[str, str]) -> None:
    keys = [key for key, value in parsed.items() if key != "notes" and str(value or "").strip()]
    if not keys:
        return
    ordered = sorted(keys)
    _append_note(parsed, f"Detected: {', '.join(ordered)}")


def _resolve_target(label: Any) -> str | None:
    normalized = _normalize_text(label)
    if not normalized:
        return None

    best_target: str | None = None
    best_score = -1

    for target, aliases in _TARGET_ALIASES.items():
        for alias in aliases:
            alias_norm = _normalize_text(alias)
            if not alias_norm:
                continue

            if normalized == alias_norm:
                score = 1000 + len(alias_norm)
            elif alias_norm in normalized:
                score = 100 + len(alias_norm)
            elif normalized in alias_norm:
                score = 50 + len(normalized)
            else:
                alias_tokens = set(alias_norm.split())
                label_tokens = set(normalized.split())
                overlap = len(alias_tokens.intersection(label_tokens))
                if overlap == 0:
                    continue
                score = overlap

            if score > best_score:
                best_score = score
                best_target = target

    return best_target


def _coerce_value(target: str, value: Any) -> str:
    text = _collapse_ws(str(value or "").strip())
    if not text:
        return ""
    lowered = text.lower()

    if target == "ic_acknowledged":
        if lowered in {"true", "yes", "y", "1", "checked"}:
            return "true"
        if lowered in {"false", "no", "n", "0", "unchecked"}:
            return "false"
    if target == "bank_account_type":
        if "check" in lowered:
            return "checking"
        if "sav" in lowered:
            return "savings"
    return text


def _row_has_any_value(row: list[Any]) -> bool:
    return any(str(cell or "").strip() for cell in row)


def _parse_table_mode_rows(
    rows: list[list[Any]],
    source_label: str,
    extra_note: str = "",
    header_scan_limit: int = 30,
) -> dict[str, str] | None:
    if not rows:
        return None

    header_idx = None
    header_map: dict[int, str] = {}
    scan_to = min(len(rows), header_scan_limit)
    for idx in range(scan_to):
        row = rows[idx]
        candidate: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            target = _resolve_target(cell)
            if target:
                candidate[col_idx] = target
        if len(candidate) >= 3:
            header_idx = idx
            header_map = candidate
            break

    if header_idx is None:
        return None

    data_row_idx = None
    for idx in range(header_idx + 1, len(rows)):
        if _row_has_any_value(rows[idx]):
            data_row_idx = idx
            break
    if data_row_idx is None:
        return None

    parsed: dict[str, str] = {}
    values = rows[data_row_idx]
    for col_idx, target in header_map.items():
        value = values[col_idx] if col_idx < len(values) else ""
        coerced = _coerce_value(target, value)
        if coerced:
            parsed[target] = coerced

    if not parsed:
        return None

    data_rows_count = len([row for row in rows[header_idx + 1 :] if _row_has_any_value(row)])
    if data_rows_count > 1 and source_label == "csv":
        _append_note(parsed, f"CSV contained {data_rows_count} rows; parsed first row.")
    elif data_rows_count > 1 and source_label == "xlsx":
        _append_note(parsed, f"XLSX contained {data_rows_count} rows; parsed first row.")

    _append_note(parsed, f"Parsed via {source_label}/table{extra_note}.")
    return parsed


def _parse_qa_mode_rows(rows: list[list[Any]], source_label: str, extra_note: str = "") -> dict[str, str] | None:
    parsed: dict[str, str] = {}
    scan_to = min(len(rows), 200)
    for idx in range(scan_to):
        row = rows[idx]
        question = row[0] if len(row) > 0 else ""
        answer = row[1] if len(row) > 1 else ""
        target = _resolve_target(question)
        if not target:
            continue
        value = _coerce_value(target, answer)
        if value:
            parsed[target] = value

    if not parsed:
        return None

    _append_note(parsed, f"Parsed via {source_label}/qa{extra_note}.")
    return parsed


def parse_onboarding_csv(file_bytes: bytes) -> dict[str, Any]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(text.splitlines())
    rows = [list(row) for row in reader]
    if not rows:
        return {}

    first_row = rows[0] if rows else []
    if len(first_row) >= 2:
        table_parsed = _parse_table_mode_rows(rows, source_label="csv")
        if table_parsed:
            _append_detected_keys_note(table_parsed)
            return table_parsed

    qa_rows: list[list[Any]] = []
    for line in text.splitlines():
        if not line.strip():
            continue
        if "," in line:
            q, a = line.split(",", 1)
        else:
            q, a = line, ""
        qa_rows.append([q, a])
    qa_parsed = _parse_qa_mode_rows(qa_rows, source_label="csv")
    if qa_parsed:
        _append_detected_keys_note(qa_parsed)
    return qa_parsed or {}


def parse_onboarding_xlsx(file_bytes: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        extra = f" on sheet '{sheet.title}'"

        table_parsed = _parse_table_mode_rows(rows, source_label="xlsx", extra_note=extra, header_scan_limit=30)
        if table_parsed:
            _append_detected_keys_note(table_parsed)
            return table_parsed

        qa_parsed = _parse_qa_mode_rows(rows, source_label="xlsx", extra_note=extra)
        if qa_parsed:
            _append_detected_keys_note(qa_parsed)
            return qa_parsed

    return {}


def parse_onboarding_file(file_bytes: bytes, filename: str) -> dict[str, Any]:
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        pdf = parse_onboarding_pdf(file_bytes)
        return {
            "name": pdf.get("name", ""),
            "business_name": pdf.get("business_name", ""),
            "email": pdf.get("email", ""),
            "phone": pdf.get("phone", ""),
            "address": pdf.get("address", ""),
            "dob": pdf.get("dob", ""),
            "ssn_or_ein": pdf.get("ein_or_ssn", ""),
            "bank_routing_number": pdf.get("bank_routing", ""),
            "bank_account_number": pdf.get("bank_account", ""),
            "notes": pdf.get("notes", ""),
        }
    if name.endswith(".csv"):
        return parse_onboarding_csv(file_bytes)
    if name.endswith(".xlsx"):
        return parse_onboarding_xlsx(file_bytes)
    raise ValueError("unsupported file type")
