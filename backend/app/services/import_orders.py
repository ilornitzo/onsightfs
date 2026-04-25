from __future__ import annotations

import csv
import hashlib
import io
import os
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.import_batch import ImportBatch
from app.models.order import Order
from app.schemas.import_orders import ImportOrderRowError, ImportOrdersResponse

HEADER_FIELD_MAP: dict[str, str | None] = {
    "ID": "order_uid",
    "OrderNumber": "order_number",
    "WorkCode": "work_code",
    "Client": "requestor_name_raw",
    "Inspector": "contractor_name_raw",
    "InspectorUserID": "contractor_user_id",
    "CountyName": "county",
    "City": "city",
    "State": "state",
    "Zip": "zip",
    "Address": "address",
    "ClientPay": "client_pay_amount",
    "InspectorPay": None,
    "Due": "due_date",
    "SubmittedToClient": "submitted_to_client_at",
    "SubmittedToClientBy": "submitted_to_client_by",
    "Submitted": "submitted_at",
    "Completed": "completed_at",
    "Assigned": "assigned_at",
    "Ordered": "ordered_at",
    "Imported": "imported_at",
    "WindowStartDate": "window_start_date",
    "WindowEndDate": "window_end_date",
    "ECD": "ecd_date",
    "DataEntryErrorCode": "data_entry_error_code",
    "Owner": "owner",
    "Lender": "lender",
    "LoanNumber": "loan_number",
    "Vacant": "vacant",
    "PhotoRequired": "photo_required",
    "Instructions": "instructions",
    "Note": "note",
    "Latitude": "latitude",
    "Longitude": "longitude",
    "QCUser": "qc_user",
    "MappingAddress1": "mapping_address_1",
    "MappingAddress2": "mapping_address_2",
    "MappingCity": "mapping_city",
    "MappingState": "mapping_state",
    "MappingZip": "mapping_zip",
    "Source": "source",
}

DATE_FIELDS = {"due_date", "inspector_due_date", "window_start_date", "window_end_date", "ecd_date"}
DATETIME_FIELDS = {
    "submitted_to_client_at",
    "submitted_at",
    "completed_at",
    "assigned_at",
    "ordered_at",
    "imported_at",
}
DECIMAL_FIELDS = {"client_pay_amount", "latitude", "longitude"}
BOOLEAN_FIELDS = {"vacant", "photo_required"}
EXCEL_SERIAL_MIN = 30000
EXCEL_SERIAL_MAX = 80000
ALTERNATE_HEADER_CANDIDATES = {
    "city",
    "state",
    "zip",
    "county",
    "countyname",
    "client",
    "inspector",
    "ordernumber",
    "order number",
    "submittedtoclient",
    "source",
    "clientpay",
    "inspectorpay",
}


def _excel_serial_to_datetime(serial: float) -> datetime:
    # Excel's 1900-date system epoch, accounting for Excel's leap-year bug behavior.
    excel_epoch = datetime(1899, 12, 30)
    return excel_epoch + timedelta(days=serial)


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.strip('"').strip("'")
    return " ".join(text.split())


def _header_lookup_key(value: str) -> str:
    return _normalize_header(value).lower()


def _normalize_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _coerce_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)

    lowered = str(value).strip().lower()
    if lowered in {"yes", "y", "true", "1"}:
        return True
    if lowered in {"no", "n", "false", "0"}:
        return False
    return None


def _coerce_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return Decimal(text.replace(",", ""))
    except InvalidOperation as exc:
        raise ValueError(f"invalid decimal value: {value}") from exc


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric_value = float(value)
        if EXCEL_SERIAL_MIN <= numeric_value <= EXCEL_SERIAL_MAX:
            return _excel_serial_to_datetime(numeric_value).date()
        return None

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"invalid date value: {value}")


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric_value = float(value)
        if EXCEL_SERIAL_MIN <= numeric_value <= EXCEL_SERIAL_MAX:
            return _excel_serial_to_datetime(numeric_value)
        return None

    text = str(value).strip()
    if not text:
        return None

    if text == "0001-01-01 00:00:00":
        return None

    for fmt in ("%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"invalid date value: {value}")


def _find_header_row(sheet: Any) -> tuple[int, dict[str, int], bool]:
    max_rows_to_scan = 50
    scanned_first_cells: list[str] = []

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_rows_to_scan, values_only=True),
        start=1,
    ):
        normalized = [_normalize_header(cell) for cell in row]
        normalized_lower = [value.lower() for value in normalized if value]

        first_non_empty = next((value for value in normalized if value), "")
        scanned_first_cells.append(f"row {row_idx}: {first_non_empty!r}")

        if "id" in normalized_lower and "address" in normalized_lower:
            header_map = {
                _header_lookup_key(header): idx
                for idx, header in enumerate(normalized)
                if header
            }
            return row_idx, header_map, False

        alternate_header_matches = len(set(normalized_lower) & ALTERNATE_HEADER_CANDIDATES)
        if "address" in normalized_lower and alternate_header_matches >= 2:
            header_map = {
                _header_lookup_key(header): idx
                for idx, header in enumerate(normalized)
                if header
            }
            return row_idx, header_map, True

    raise ValueError(
        f"header row not found after scanning {max_rows_to_scan} rows; "
        f"first non-empty cells: {', '.join(scanned_first_cells)}"
    )


def _get_by_header(row: tuple[Any, ...], header_map: dict[str, int], header_name: str) -> Any:
    idx = header_map.get(_header_lookup_key(header_name))
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _coerce_field(field_name: str, value: Any) -> Any:
    if field_name in BOOLEAN_FIELDS:
        return _coerce_bool(value)
    if field_name in DECIMAL_FIELDS:
        return _coerce_decimal(value)
    if field_name in DATE_FIELDS:
        return parse_date(value)
    if field_name in DATETIME_FIELDS:
        return parse_datetime(value)
    return _normalize_string(value)


def _detect_file_extension(filename: str | None) -> str:
    if not filename:
        return ""
    return os.path.splitext(filename)[1].lower()


def _load_rows_from_xlsx(
    file_obj: BinaryIO,
) -> tuple[list[tuple[int, tuple[Any, ...]]], dict[str, int], bool]:
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_row_idx, header_map, is_alternate_header = _find_header_row(sheet)
        rows: list[tuple[int, tuple[Any, ...]]] = []
        for excel_row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            if row is None or not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue
            rows.append((excel_row_idx, row))
        return rows, header_map, is_alternate_header
    finally:
        workbook.close()


def _build_generated_order_uid(
    *,
    order_number: str | None,
    address: str | None,
    city: str | None,
    state: str | None,
    zip_code: str | None,
    submitted_to_client_at: datetime | None,
) -> str:
    if order_number:
        seed_parts = ["order_number", order_number.strip()]
    else:
        submitted_value = submitted_to_client_at.isoformat() if submitted_to_client_at else ""
        seed_parts = [
            "address_bundle",
            address or "",
            city or "",
            state or "",
            zip_code or "",
            submitted_value,
        ]

    digest = hashlib.sha256("|".join(seed_parts).encode("utf-8")).hexdigest()
    return f"generated-{digest}"


def _load_rows_from_csv(file_obj: BinaryIO) -> tuple[list[tuple[int, tuple[Any, ...]]], dict[str, int]]:
    raw_content = file_obj.read()
    if isinstance(raw_content, bytes):
        text = raw_content.decode("utf-8-sig", errors="replace")
    else:
        text = str(raw_content)

    reader = csv.reader(io.StringIO(text))
    header_map: dict[str, int] | None = None
    rows: list[tuple[int, tuple[Any, ...]]] = []

    for csv_row_idx, row in enumerate(reader, start=1):
        normalized = [_normalize_header(cell) for cell in row]
        is_non_empty = any(value for value in normalized)

        if header_map is None:
            if not is_non_empty:
                continue
            header_map = {
                _header_lookup_key(header): idx
                for idx, header in enumerate(normalized)
                if header
            }
            continue

        rows.append((csv_row_idx, tuple(row)))

    if header_map is None:
        raise ValueError("header row not found in CSV (first non-empty row expected as header)")

    return rows, header_map


def import_orders_file(db: Session, file_obj: BinaryIO, filename: str | None) -> ImportOrdersResponse:
    extension = _detect_file_extension(filename)
    is_alternate_header = False
    file_obj.seek(0)
    if extension == ".csv":
        source_rows, header_map = _load_rows_from_csv(file_obj)
    elif extension in {".xlsx", ".xls"}:
        source_rows, header_map, is_alternate_header = _load_rows_from_xlsx(file_obj)
    else:
        raise ValueError("unsupported file type")

    try:
        import_batch = ImportBatch(
            source_system="InspectorADE",
            source_file_name=filename or "",
            row_count=0,
            inserted_count=0,
            duplicate_count=0,
            error_count=0,
        )
        db.add(import_batch)
        db.flush()

        row_count = 0
        inserted_count = 0
        duplicate_count = 0
        error_count = 0
        errors: list[ImportOrderRowError] = []

        for excel_row_idx, row in source_rows:
            if row is None or not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue

            row_count += 1
            order_uid_value = _get_by_header(row, header_map, "ID")
            order_uid = _normalize_string(order_uid_value)

            try:
                order_data: dict[str, Any] = {
                    "import_batch_id": import_batch.id,
                    "client_name_raw": _coerce_field(
                        "client_name_raw",
                        _get_by_header(row, header_map, "Source"),
                    ),
                }

                for header_name, field_name in HEADER_FIELD_MAP.items():
                    if field_name is None or field_name == "order_uid":
                        continue
                    if field_name == "requestor_name_raw" and not hasattr(Order, "requestor_name_raw"):
                        continue
                    raw_value = _get_by_header(row, header_map, header_name)
                    order_data[field_name] = _coerce_field(field_name, raw_value)

                order_number = (
                    _get_by_header(row, header_map, "OrderNumber")
                    or _get_by_header(row, header_map, "order_number")
                    or _get_by_header(row, header_map, "Order Number")
                )
                order_data["order_number"] = str(order_number).strip() if order_number else None
                if not order_uid and is_alternate_header:
                    order_uid = _build_generated_order_uid(
                        order_number=order_data["order_number"],
                        address=order_data.get("address"),
                        city=order_data.get("city"),
                        state=order_data.get("state"),
                        zip_code=order_data.get("zip"),
                        submitted_to_client_at=order_data.get("submitted_to_client_at"),
                    )

                if not order_uid:
                    error_count += 1
                    errors.append(
                        ImportOrderRowError(
                            row=excel_row_idx,
                            order_uid=None,
                            message="missing required field: ID",
                        )
                    )
                    continue

                order_data["order_uid"] = order_uid

                existing = None
                if order_data.get("order_number"):
                    existing = db.query(Order).filter(Order.order_number == order_data["order_number"]).first()
                else:
                    existing = db.query(Order).filter(Order.order_uid == order_uid).first()

                if existing:
                    duplicate_count += 1
                    continue

                db.add(Order(**order_data))
                inserted_count += 1
            except Exception as exc:
                error_count += 1
                errors.append(
                    ImportOrderRowError(
                        row=excel_row_idx,
                        order_uid=order_uid,
                        message=str(exc),
                    )
                )

        import_batch.row_count = row_count
        import_batch.inserted_count = inserted_count
        import_batch.duplicate_count = duplicate_count
        import_batch.error_count = error_count

        db.commit()
        return ImportOrdersResponse(
            import_batch_id=str(import_batch.id),
            row_count=row_count,
            inserted_count=inserted_count,
            duplicate_count=duplicate_count,
            error_count=error_count,
            errors=errors,
        )
    except Exception:
        db.rollback()
        raise
